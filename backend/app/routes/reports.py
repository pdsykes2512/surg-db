"""
Report generation API routes
Updated to work with cancer episodes and treatments in separate collections
"""
from fastapi import APIRouter, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from typing import Dict, Any, Optional
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..database import get_surgeries_collection, get_episodes_collection, get_treatments_collection, get_tumours_collection


router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/summary")
async def get_summary_report() -> Dict[str, Any]:
    """Get overall outcome statistics from surgical treatments in separate collection"""
    treatments_collection = await get_treatments_collection()
    episodes_collection = await get_episodes_collection()
    
    # Query treatments collection directly (no unwinding needed)
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {
            "$facet": {
                "total": [{"$count": "count"}],
                "urgency": [
                    {"$group": {"_id": "$urgency", "count": {"$sum": 1}}}
                ],
                "avg_duration": [
                    {"$match": {"operation_duration_minutes": {"$exists": True, "$ne": None}}},
                    {"$group": {"_id": None, "avg": {"$avg": "$operation_duration_minutes"}}}
                ],
                "complications": [
                    {"$match": {"clavien_dindo_grade": {"$exists": True, "$ne": None}}},
                    {"$count": "count"}
                ],
                "readmissions": [
                    {"$match": {"readmission_30d": True}},
                    {"$count": "count"}
                ],
                "return_theatre": [
                    {"$match": {"return_to_theatre": True}},
                    {"$count": "count"}
                ],
                "avg_los": [
                    {"$match": {"length_of_stay": {"$exists": True, "$ne": None}}},
                    {"$group": {"_id": None, "avg": {"$avg": "$length_of_stay"}}}
                ]
            }
        }
    ]
    
    result = await treatments_collection.aggregate(pipeline).to_list(length=1)
    if not result or not result[0]["total"]:
        return {"total_surgeries": 0, "complication_rate": 0, "readmission_rate": 0, "mortality_rate": 0, 
                "return_to_theatre_rate": 0, "escalation_rate": 0, "avg_length_of_stay_days": 0, 
                "urgency_breakdown": {}, "generated_at": datetime.utcnow().isoformat()}
    
    stats = result[0]
    total_surgeries = stats["total"][0]["count"] if stats["total"] else 0
    
    # Calculate rates
    complications_count = stats["complications"][0]["count"] if stats["complications"] else 0
    readmissions_count = stats["readmissions"][0]["count"] if stats["readmissions"] else 0
    return_theatre_count = stats["return_theatre"][0]["count"] if stats["return_theatre"] else 0
    
    complication_rate = complications_count / total_surgeries if total_surgeries > 0 else 0
    readmission_rate = readmissions_count / total_surgeries if total_surgeries > 0 else 0
    return_to_theatre_rate = return_theatre_count / total_surgeries if total_surgeries > 0 else 0
    
    # Get mortality from episodes (death within 30 days of treatment)
    mortality_count = await episodes_collection.count_documents({
        "outcome.patient_status": "deceased",
        "outcome.date_of_death": {"$exists": True, "$ne": None}
    })
    mortality_rate = mortality_count / total_surgeries if total_surgeries > 0 else 0
    
    # Escalation rate (ICU escalation) - would need to be added to treatment model
    escalation_rate = 0  # Placeholder until we add this field
    
    urgency_breakdown = {item["_id"] if item["_id"] else "unknown": item["count"] for item in stats["urgency"]}
    # Ensure all urgencies are present
    for urgency in ["elective", "emergency", "urgent"]:
        if urgency not in urgency_breakdown:
            urgency_breakdown[urgency] = 0
    
    avg_duration = round(stats["avg_duration"][0]["avg"], 2) if stats["avg_duration"] and stats["avg_duration"][0] else 0
    avg_los = round(stats["avg_los"][0]["avg"], 2) if stats["avg_los"] and stats["avg_los"][0] else 0
    
    return {
        "total_surgeries": total_surgeries,
        "avg_operation_duration_minutes": avg_duration,
        "complication_rate": complication_rate,
        "readmission_rate": readmission_rate,
        "mortality_rate": mortality_rate,
        "return_to_theatre_rate": return_to_theatre_rate,
        "escalation_rate": escalation_rate,
        "avg_length_of_stay_days": avg_los,
        "urgency_breakdown": urgency_breakdown,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/complications")
async def get_complications_report() -> Dict[str, Any]:
    """Get detailed complication analysis from surgical treatments"""
    # Currently no complications data in simplified structure
    # This will need to be populated once full treatment models are used
    
    return {
        "complications_by_type": [],
        "message": "Complications tracking not yet implemented in current data structure",
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/trends")
async def get_trends_report(
    days: int = Query(30, description="Number of days to analyze", ge=1, le=365)
) -> Dict[str, Any]:
    """Get trends over specified time period from surgical treatments in separate collection"""
    treatments_collection = await get_treatments_collection()
    
    start_date = datetime.utcnow() - timedelta(days=days)
    
    # Surgeries by date from treatments collection
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {"$match": {"treatment_date": {"$exists": True}}},
        {"$addFields": {
            "treatment_date_parsed": {
                "$dateFromString": {
                    "dateString": "$treatment_date",
                    "onError": None,
                    "onNull": None
                }
            }
        }},
        {"$match": {"treatment_date_parsed": {"$gte": start_date, "$ne": None}}},
        {"$group": {
            "_id": {"$dateToString": {"format": "%Y-%m-%d", "date": "$treatment_date_parsed"}},
            "count": {"$sum": 1}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    daily_trends = await treatments_collection.aggregate(pipeline).to_list(length=days)
    
    return {
        "period_days": days,
        "start_date": start_date.isoformat(),
        "end_date": datetime.utcnow().isoformat(),
        "daily_trends": daily_trends,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/surgeon-performance")
async def get_surgeon_performance() -> Dict[str, Any]:
    """Get surgeon-specific performance metrics from surgical treatments in separate collection"""
    treatments_collection = await get_treatments_collection()
    
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {"$group": {
            "_id": "$surgeon",
            "total_surgeries": {"$sum": 1},
            "avg_duration": {"$avg": "$operation_duration_minutes"},
            "complications": {
                "$sum": {
                    "$cond": [
                        {"$and": [
                            {"$ne": [{"$ifNull": ["$clavien_dindo_grade", None]}, None]},
                            {"$ne": [{"$ifNull": ["$clavien_dindo_grade", ""]}, ""]}
                        ]}, 
                        1, 
                        0
                    ]
                }
            },
            "readmissions": {
                "$sum": {"$cond": [{"$eq": ["$readmission_30d", True]}, 1, 0]}
            },
            "avg_los": {"$avg": "$length_of_stay"}
        }},
        {"$sort": {"total_surgeries": -1}}
    ]
    
    surgeon_stats = await treatments_collection.aggregate(pipeline).to_list(length=100)
    
    # Calculate rates and round numeric values
    for stat in surgeon_stats:
        if stat["avg_duration"]:
            stat["avg_duration"] = round(stat["avg_duration"], 2)
        if stat["avg_los"]:
            stat["avg_los"] = round(stat["avg_los"], 2)
        else:
            stat["avg_los"] = 0.0
        
        # Calculate complication and readmission rates
        total = stat["total_surgeries"]
        stat["complication_rate"] = stat["complications"] / total if total > 0 else 0.0
        stat["readmission_rate"] = stat["readmissions"] / total if total > 0 else 0.0
        
        # Remove intermediate counts (keep only rates)
        del stat["complications"]
        del stat["readmissions"]
    
    return {
        "surgeons": surgeon_stats,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/export/summary-excel")
async def export_summary_excel():
    """Export summary report to Excel"""
    # Get the summary data
    summary_data = await get_summary_report()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary Report"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Title
    ws['A1'] = "Surgical Outcomes Summary Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    ws['A2'] = f"Generated: {summary_data['generated_at']}"
    ws.merge_cells('A2:B2')
    
    # Overall Statistics
    row = 4
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    
    stats = [
        ("Total Surgeries", summary_data['total_surgeries']),
        ("Avg Operation Duration (min)", summary_data['avg_operation_duration_minutes']),
        ("Complication Rate (%)", f"{summary_data['complication_rate']*100:.1f}"),
        ("Readmission Rate (%)", f"{summary_data['readmission_rate']*100:.1f}"),
        ("Mortality Rate (%)", f"{summary_data['mortality_rate']*100:.1f}"),
        ("Return to Theatre Rate (%)", f"{summary_data['return_to_theatre_rate']*100:.1f}"),
        ("Escalation Rate (%)", f"{summary_data['escalation_rate']*100:.1f}"),
        ("Avg Length of Stay (days)", summary_data['avg_length_of_stay_days']),
    ]
    
    row += 1
    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        row += 1
    
    # Urgency Breakdown
    row += 1
    ws[f'A{row}'] = "Urgency Breakdown"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    
    row += 1
    for urgency, count in summary_data['urgency_breakdown'].items():
        ws[f'A{row}'] = urgency.capitalize()
        ws[f'B{row}'] = count
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"surgical_summary_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/surgeon-performance-excel")
async def export_surgeon_performance_excel():
    """Export surgeon performance report to Excel"""
    # Get the surgeon performance data
    perf_data = await get_surgeon_performance()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Surgeon Performance"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = "Surgeon Performance Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = f"Generated: {perf_data['generated_at']}"
    ws.merge_cells('A2:F2')
    
    # Headers
    row = 4
    headers = ["Surgeon", "Total Surgeries", "Avg Duration (min)", "Complication Rate (%)", "Readmission Rate (%)", "Avg LOS (days)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
    
    # Data rows
    row += 1
    for surgeon in perf_data['surgeons']:
        ws.cell(row=row, column=1, value=surgeon['_id'] or "Unknown")
        ws.cell(row=row, column=2, value=surgeon['total_surgeries'])
        ws.cell(row=row, column=3, value=surgeon['avg_duration'] or 0)
        ws.cell(row=row, column=4, value=f"{surgeon['complication_rate']*100:.1f}")
        ws.cell(row=row, column=5, value=f"{surgeon['readmission_rate']*100:.1f}")
        ws.cell(row=row, column=6, value=surgeon['avg_los'] or 0)
        
        # Center align numeric columns
        for col in range(2, 7):
            ws.cell(row=row, column=col).alignment = center_aligned
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"surgeon_performance_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/trends-excel")
async def export_trends_excel(days: int = Query(30, description="Number of days to analyze", ge=1, le=365)):
    """Export trends report to Excel"""
    # Get the trends data
    trends_data = await get_trends_report(days=days)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Surgery Trends"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = f"Surgery Trends - Last {days} Days"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    
    ws['A2'] = f"Period: {trends_data['start_date'][:10]} to {trends_data['end_date'][:10]}"
    ws.merge_cells('A2:B2')
    
    ws['A3'] = f"Generated: {trends_data['generated_at']}"
    ws.merge_cells('A3:B3')
    
    # Headers
    row = 5
    ws.cell(row=row, column=1, value="Date").font = header_font
    ws.cell(row=row, column=1).fill = header_fill
    ws.cell(row=row, column=2, value="Number of Surgeries").font = header_font
    ws.cell(row=row, column=2).fill = header_fill
    
    # Data rows
    row += 1
    total_surgeries = 0
    for trend in trends_data['daily_trends']:
        ws.cell(row=row, column=1, value=trend['_id'])
        ws.cell(row=row, column=2, value=trend['count'])
        ws.cell(row=row, column=2).alignment = center_aligned
        total_surgeries += trend['count']
        row += 1
    
    # Summary
    row += 1
    ws.cell(row=row, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=row, column=2, value=total_surgeries).font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = center_aligned
    
    row += 1
    ws.cell(row=row, column=1, value="Daily Average").font = Font(bold=True)
    avg = total_surgeries / days if days > 0 else 0
    ws.cell(row=row, column=2, value=round(avg, 2)).font = Font(bold=True)
    ws.cell(row=row, column=2).alignment = center_aligned
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"surgery_trends_{days}days_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============== NBOCA-Specific Reports ==============

@router.get("/nboca/mortality")
async def get_nboca_mortality_report() -> Dict[str, Any]:
    """Get 30-day and 90-day mortality rates for NBOCA reporting"""
    treatments_collection = await get_treatments_collection()
    
    # Get all surgery treatments
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {
            "$facet": {
                "total": [{"$count": "count"}],
                "mortality_30day": [
                    {"$match": {"outcomes.mortality_30day": True}},
                    {"$count": "count"}
                ],
                "mortality_90day": [
                    {"$match": {"outcomes.mortality_90day": True}},
                    {"$count": "count"}
                ],
                "with_urgency": [
                    {"$group": {
                        "_id": "$urgency",
                        "total": {"$sum": 1},
                        "mortality_30day": {
                            "$sum": {"$cond": [{"$eq": ["$outcomes.mortality_30day", True]}, 1, 0]}
                        },
                        "mortality_90day": {
                            "$sum": {"$cond": [{"$eq": ["$outcomes.mortality_90day", True]}, 1, 0]}
                        }
                    }}
                ]
            }
        }
    ]
    
    result = await treatments_collection.aggregate(pipeline).to_list(length=1)
    
    if not result or not result[0]["total"]:
        return {
            "total_surgeries": 0,
            "mortality_30day_count": 0,
            "mortality_30day_rate": 0.0,
            "mortality_90day_count": 0,
            "mortality_90day_rate": 0.0,
            "by_urgency": [],
            "generated_at": datetime.utcnow().isoformat()
        }
    
    stats = result[0]
    total = stats["total"][0]["count"] if stats["total"] else 0
    mort_30 = stats["mortality_30day"][0]["count"] if stats["mortality_30day"] else 0
    mort_90 = stats["mortality_90day"][0]["count"] if stats["mortality_90day"] else 0
    
    # Calculate rates by urgency
    by_urgency = []
    for urg_stat in stats["with_urgency"]:
        urgency = urg_stat["_id"] or "unknown"
        urg_total = urg_stat["total"]
        urg_mort_30 = urg_stat["mortality_30day"]
        urg_mort_90 = urg_stat["mortality_90day"]
        
        by_urgency.append({
            "urgency": urgency,
            "total_surgeries": urg_total,
            "mortality_30day_count": urg_mort_30,
            "mortality_30day_rate": round((urg_mort_30 / urg_total * 100), 2) if urg_total > 0 else 0.0,
            "mortality_90day_count": urg_mort_90,
            "mortality_90day_rate": round((urg_mort_90 / urg_total * 100), 2) if urg_total > 0 else 0.0,
        })
    
    return {
        "total_surgeries": total,
        "mortality_30day_count": mort_30,
        "mortality_30day_rate": round((mort_30 / total * 100), 2) if total > 0 else 0.0,
        "mortality_90day_count": mort_90,
        "mortality_90day_rate": round((mort_90 / total * 100), 2) if total > 0 else 0.0,
        "by_urgency": by_urgency,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/nboca/anastomotic-leak")
async def get_nboca_anastomotic_leak_report() -> Dict[str, Any]:
    """Get anastomotic leak rates for NBOCA colorectal surgery reporting"""
    treatments_collection = await get_treatments_collection()
    
    # Get all colorectal surgeries (bowel resections with anastomosis)
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {"$match": {"procedure.opcs4_codes": {"$exists": True}}},
        {
            "$facet": {
                "total_anastomoses": [
                    # Filter for anastomosis procedures (primary anastomosis or defunctioning)
                    {"$match": {
                        "$or": [
                            {"intraoperative.anastomosis.performed": True},
                            {"procedure.opcs4_codes": {"$regex": "^(H04|H06|H07|H08|H33)"}}
                        ]
                    }},
                    {"$count": "count"}
                ],
                "with_leak": [
                    {"$match": {"postoperative_events.complications": {"$exists": True, "$ne": []}}},
                    {"$unwind": "$postoperative_events.complications"},
                    {"$match": {
                        "postoperative_events.complications.type": {
                            "$regex": "anastomotic.*leak|leak.*anastomosis|AL", 
                            "$options": "i"
                        }
                    }},
                    {"$group": {"_id": "$treatment_id"}},
                    {"$count": "count"}
                ],
                "by_procedure": [
                    {"$match": {"intraoperative.anastomosis.performed": True}},
                    {"$group": {
                        "_id": "$procedure.primary_procedure",
                        "total": {"$sum": 1},
                        "leaks": {
                            "$sum": {
                                "$cond": [
                                    {
                                        "$anyElementTrue": {
                                            "$map": {
                                                "input": {"$ifNull": ["$postoperative_events.complications", []]},
                                                "as": "comp",
                                                "in": {
                                                    "$regexMatch": {
                                                        "input": "$$comp.type",
                                                        "regex": "anastomotic.*leak|leak.*anastomosis|AL",
                                                        "options": "i"
                                                    }
                                                }
                                            }
                                        }
                                    },
                                    1,
                                    0
                                ]
                            }
                        }
                    }}
                ]
            }
        }
    ]
    
    result = await treatments_collection.aggregate(pipeline).to_list(length=1)
    
    if not result:
        return {
            "total_anastomoses": 0,
            "leak_count": 0,
            "leak_rate": 0.0,
            "by_procedure": [],
            "generated_at": datetime.utcnow().isoformat()
        }
    
    stats = result[0]
    total = stats["total_anastomoses"][0]["count"] if stats["total_anastomoses"] else 0
    leaks = stats["with_leak"][0]["count"] if stats["with_leak"] else 0
    
    # Format by procedure
    by_procedure = []
    for proc in stats["by_procedure"]:
        proc_total = proc["total"]
        proc_leaks = proc["leaks"]
        by_procedure.append({
            "procedure": proc["_id"] or "Unknown",
            "total_anastomoses": proc_total,
            "leak_count": proc_leaks,
            "leak_rate": round((proc_leaks / proc_total * 100), 2) if proc_total > 0 else 0.0
        })
    
    return {
        "total_anastomoses": total,
        "leak_count": leaks,
        "leak_rate": round((leaks / total * 100), 2) if total > 0 else 0.0,
        "by_procedure": by_procedure,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/nboca/conversion-rates")
async def get_nboca_conversion_rates() -> Dict[str, Any]:
    """Get laparoscopic to open conversion rates for NBOCA reporting"""
    treatments_collection = await get_treatments_collection()
    
    pipeline = [
        {"$match": {"treatment_type": "surgery"}},
        {"$match": {"intraoperative.surgical_approach": {"$exists": True}}},
        {
            "$facet": {
                "laparoscopic_started": [
                    {"$match": {"intraoperative.surgical_approach": {"$regex": "laparoscopic", "$options": "i"}}},
                    {"$count": "count"}
                ],
                "conversions": [
                    {"$match": {"intraoperative.conversion_to_open": True}},
                    {"$count": "count"}
                ],
                "by_procedure": [
                    {"$match": {"intraoperative.surgical_approach": {"$regex": "laparoscopic", "$options": "i"}}},
                    {"$group": {
                        "_id": "$procedure.primary_procedure",
                        "total_laparoscopic": {"$sum": 1},
                        "conversions": {
                            "$sum": {"$cond": [{"$eq": ["$intraoperative.conversion_to_open", True]}, 1, 0]}
                        }
                    }}
                ]
            }
        }
    ]
    
    result = await treatments_collection.aggregate(pipeline).to_list(length=1)
    
    if not result:
        return {
            "total_laparoscopic": 0,
            "conversion_count": 0,
            "conversion_rate": 0.0,
            "by_procedure": [],
            "generated_at": datetime.utcnow().isoformat()
        }
    
    stats = result[0]
    total_lap = stats["laparoscopic_started"][0]["count"] if stats["laparoscopic_started"] else 0
    conversions = stats["conversions"][0]["count"] if stats["conversions"] else 0
    
    # Format by procedure
    by_procedure = []
    for proc in stats["by_procedure"]:
        proc_total = proc["total_laparoscopic"]
        proc_conv = proc["conversions"]
        by_procedure.append({
            "procedure": proc["_id"] or "Unknown",
            "total_laparoscopic": proc_total,
            "conversion_count": proc_conv,
            "conversion_rate": round((proc_conv / proc_total * 100), 2) if proc_total > 0 else 0.0
        })
    
    return {
        "total_laparoscopic": total_lap,
        "conversion_count": conversions,
        "conversion_rate": round((conversions / total_lap * 100), 2) if total_lap > 0 else 0.0,
        "by_procedure": by_procedure,
        "generated_at": datetime.utcnow().isoformat()
    }


@router.get("/export/nboca-mortality-excel")
async def export_nboca_mortality_excel():
    """Export NBOCA mortality report to Excel"""
    # Get the mortality data
    mortality_data = await get_nboca_mortality_report()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "NBOCA Mortality"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = "NBOCA 30-Day and 90-Day Mortality Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {mortality_data['generated_at']}"
    ws.merge_cells('A2:D2')
    
    # Overall Stats
    row = 4
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Total Surgeries"
    ws[f'B{row}'] = mortality_data['total_surgeries']
    
    row += 1
    ws[f'A{row}'] = "30-Day Mortality Count"
    ws[f'B{row}'] = mortality_data['mortality_30day_count']
    
    row += 1
    ws[f'A{row}'] = "30-Day Mortality Rate (%)"
    ws[f'B{row}'] = mortality_data['mortality_30day_rate']
    
    row += 1
    ws[f'A{row}'] = "90-Day Mortality Count"
    ws[f'B{row}'] = mortality_data['mortality_90day_count']
    
    row += 1
    ws[f'A{row}'] = "90-Day Mortality Rate (%)"
    ws[f'B{row}'] = mortality_data['mortality_90day_rate']
    
    # By Urgency
    row += 2
    ws[f'A{row}'] = "Mortality by Urgency"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    headers = ["Urgency", "Total", "30-Day Deaths", "30-Day Rate (%)", "90-Day Deaths", "90-Day Rate (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_aligned
    
    row += 1
    for urg in mortality_data['by_urgency']:
        ws.cell(row=row, column=1, value=urg['urgency'].capitalize())
        ws.cell(row=row, column=2, value=urg['total_surgeries'])
        ws.cell(row=row, column=3, value=urg['mortality_30day_count'])
        ws.cell(row=row, column=4, value=urg['mortality_30day_rate'])
        ws.cell(row=row, column=5, value=urg['mortality_90day_count'])
        ws.cell(row=row, column=6, value=urg['mortality_90day_rate'])
        
        for col in range(2, 7):
            ws.cell(row=row, column=col).alignment = center_aligned
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 18
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"nboca_mortality_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/nboca-anastomotic-leak-excel")
async def export_nboca_anastomotic_leak_excel():
    """Export NBOCA anastomotic leak report to Excel"""
    # Get the leak data
    leak_data = await get_nboca_anastomotic_leak_report()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Anastomotic Leak"
    
    # Styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_aligned = Alignment(horizontal="center")
    
    # Title
    ws['A1'] = "NBOCA Anastomotic Leak Rate Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    
    ws['A2'] = f"Generated: {leak_data['generated_at']}"
    ws.merge_cells('A2:D2')
    
    # Overall Stats
    row = 4
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    
    row += 1
    ws[f'A{row}'] = "Total Anastomoses"
    ws[f'B{row}'] = leak_data['total_anastomoses']
    
    row += 1
    ws[f'A{row}'] = "Anastomotic Leak Count"
    ws[f'B{row}'] = leak_data['leak_count']
    
    row += 1
    ws[f'A{row}'] = "Anastomotic Leak Rate (%)"
    ws[f'B{row}'] = leak_data['leak_rate']
    
    # By Procedure
    if leak_data['by_procedure']:
        row += 2
        ws[f'A{row}'] = "Leak Rate by Procedure"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ["Procedure", "Total Anastomoses", "Leak Count", "Leak Rate (%)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
        
        row += 1
        for proc in leak_data['by_procedure']:
            ws.cell(row=row, column=1, value=proc['procedure'])
            ws.cell(row=row, column=2, value=proc['total_anastomoses'])
            ws.cell(row=row, column=3, value=proc['leak_count'])
            ws.cell(row=row, column=4, value=proc['leak_rate'])
            
            for col in range(2, 5):
                ws.cell(row=row, column=col).alignment = center_aligned
            row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 30
    for col in ['B', 'C', 'D']:
        ws.column_dimensions[col].width = 20
    
    # Save to bytes
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"nboca_anastomotic_leak_{timestamp}.xlsx"
    
    return StreamingResponse(
        excel_buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@router.get("/data-quality")
async def get_data_quality_report() -> Dict[str, Any]:
    """
    Get data quality metrics for NBOCA COSD compliance
    Shows completeness of critical fields across episodes and treatments
    """
    episodes_collection = await get_episodes_collection()
    treatments_collection = await get_treatments_collection()
    tumours_collection = await get_tumours_collection()
    
    # Get total counts
    total_episodes = await episodes_collection.count_documents({"condition_type": "cancer", "cancer_type": "bowel"})
    total_treatments = await treatments_collection.count_documents({"treatment_type": "surgery"})
    
    if total_episodes == 0:
        return {
            "total_episodes": 0,
            "total_treatments": 0,
            "episode_fields": [],
            "treatment_fields": [],
            "overall_completeness": 0,
            "categories": [],
            "generated_at": datetime.utcnow().isoformat()
        }
    
    # Get tumours collection for TNM data (stored separately)
    
    # Define NBOCA COSD critical fields for episodes
    episode_fields = [
        {"field": "referral_date", "label": "Referral Date", "category": "Process"},
        {"field": "referral_source", "label": "Referral Source (CR1600)", "category": "Process"},
        {"field": "provider_first_seen", "label": "Provider First Seen (CR1410)", "category": "Process"},
        {"field": "cns_involved", "label": "CNS Involved (CR2050)", "category": "Process"},
        {"field": "mdt_discussion_date", "label": "MDT Discussion Date", "category": "Process"},
        {"field": "mdt_meeting_type", "label": "MDT Meeting Type (CR3190)", "category": "Process"},
        {"field": "performance_status", "label": "Performance Status (CR0510)", "category": "Clinical"},
        {"field": "cancer_data.diagnosis_date", "label": "Diagnosis Date (CR2030)", "category": "Clinical"},
        {"field": "cancer_data.icd10_code", "label": "ICD-10 Code (CR0370)", "category": "Clinical"},
    ]
    
    # TNM fields - stored in separate tumours collection
    tnm_fields = [
        {"field": "tnm_version", "label": "TNM Version (CR2070)", "category": "Clinical"},
        {"field": "clinical_t", "label": "Clinical T Stage", "category": "Clinical"},
        {"field": "clinical_n", "label": "Clinical N Stage", "category": "Clinical"},
        {"field": "clinical_m", "label": "Clinical M Stage", "category": "Clinical"},
    ]
    
    # Define NBOCA COSD critical fields for treatments/surgeries
    treatment_fields = [
        {"field": "treatment_date", "label": "Surgery Date", "category": "Process"},
        {"field": "provider_organisation", "label": "Provider Organisation (CR1450)", "category": "Process"},
        {"field": "opcs4_code", "label": "OPCS-4 Code", "category": "Process"},
        {"field": "approach", "label": "Surgical Approach", "category": "Clinical"},
        {"field": "urgency", "label": "Urgency", "category": "Clinical"},
        {"field": "asa_score", "label": "ASA Score (CR6010)", "category": "Clinical"},
        {"field": "admission_date", "label": "Admission Date", "category": "Process"},
        {"field": "discharge_date", "label": "Discharge Date", "category": "Process"},
        {"field": "length_of_stay", "label": "Length of Stay", "category": "Outcome"},
        {"field": "clavien_dindo_grade", "label": "Clavien-Dindo Grade", "category": "Outcome"},
        {"field": "return_to_theatre", "label": "Return to Theatre", "category": "Outcome"},
        {"field": "readmission_30d", "label": "30-Day Readmission", "category": "Outcome"},
    ]
    
    # Calculate completeness for episode fields
    episode_stats = []
    for field_def in episode_fields:
        field_path = field_def["field"]
        
        # Build query to check if field exists and is not null/empty
        query = {
            "condition_type": "cancer",
            "cancer_type": "bowel",
            field_path: {"$exists": True, "$ne": None, "$ne": ""}
        }
        
        complete_count = await episodes_collection.count_documents(query)
        completeness = (complete_count / total_episodes * 100) if total_episodes > 0 else 0
        
        episode_stats.append({
            "field": field_def["label"],
            "category": field_def["category"],
            "complete_count": complete_count,
            "total_count": total_episodes,
            "completeness": round(completeness, 1),
            "missing_count": total_episodes - complete_count
        })
    
    # Calculate completeness for TNM fields from tumours collection
    for field_def in tnm_fields:
        field_path = field_def["field"]
        
        # Count tumours with this field populated
        query = {
            field_path: {"$exists": True, "$ne": None, "$ne": ""}
        }
        
        complete_count = await tumours_collection.count_documents(query)
        completeness = (complete_count / total_episodes * 100) if total_episodes > 0 else 0
        
        episode_stats.append({
            "field": field_def["label"],
            "category": field_def["category"],
            "complete_count": complete_count,
            "total_count": total_episodes,
            "completeness": round(completeness, 1),
            "missing_count": total_episodes - complete_count
        })
    
    # Calculate completeness for treatment fields
    treatment_stats = []
    for field_def in treatment_fields:
        field_path = field_def["field"]
        
        query = {
            "treatment_type": "surgery",
            field_path: {"$exists": True, "$ne": None, "$ne": ""}
        }
        
        complete_count = await treatments_collection.count_documents(query)
        completeness = (complete_count / total_treatments * 100) if total_treatments > 0 else 0
        
        treatment_stats.append({
            "field": field_def["label"],
            "category": field_def["category"],
            "complete_count": complete_count,
            "total_count": total_treatments,
            "completeness": round(completeness, 1),
            "missing_count": total_treatments - complete_count
        })
    
    # Calculate overall completeness
    all_fields = episode_stats + treatment_stats
    overall_completeness = sum(f["completeness"] for f in all_fields) / len(all_fields) if all_fields else 0
    
    # Group by category
    categories = {}
    for stat in all_fields:
        cat = stat["category"]
        if cat not in categories:
            categories[cat] = {
                "name": cat,
                "total_fields": 0,
                "avg_completeness": 0,
                "fields": []
            }
        categories[cat]["total_fields"] += 1
        categories[cat]["fields"].append(stat)
    
    # Calculate average completeness per category
    for cat_data in categories.values():
        cat_data["avg_completeness"] = round(
            sum(f["completeness"] for f in cat_data["fields"]) / len(cat_data["fields"]),
            1
        )
    
    return {
        "total_episodes": total_episodes,
        "total_treatments": total_treatments,
        "overall_completeness": round(overall_completeness, 1),
        "categories": list(categories.values()),
        "episode_fields": episode_stats,
        "treatment_fields": treatment_stats,
        "generated_at": datetime.utcnow().isoformat()
    }
